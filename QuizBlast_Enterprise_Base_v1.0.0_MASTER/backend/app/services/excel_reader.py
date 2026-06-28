from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def read_excel_rows_from_bytes(content: bytes) -> list[dict[str, Any]]:
    '''
    Reads first worksheet from xlsx content and returns list of dict rows.

    The first row is treated as header.
    Empty rows are skipped.
    '''
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [
        str(cell).strip() if cell is not None else ""
        for cell in rows[0]
    ]

    result: list[dict[str, Any]] = []

    for data_row in rows[1:]:
        row_dict = {}

        has_any_value = False

        for index, header in enumerate(headers):
            if not header:
                continue

            value = data_row[index] if index < len(data_row) else None

            if value is not None and str(value).strip() != "":
                has_any_value = True

            row_dict[header] = value

        if has_any_value:
            result.append(row_dict)

    return result
